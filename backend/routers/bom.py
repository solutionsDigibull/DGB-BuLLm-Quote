from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from database import get_db
from models.models import Project, BomLine, BomDetail, FgCount
from services.bom_parser import parse_scrub_bom
import io
import json
import difflib

router = APIRouter()

SCRUB_BOM_REQUIRED = [
    "FG", "Level", "Assembly", "CPN", "Description", "MFR", "MPN",
    "Commodity", "Quantity", "UOM", "Part Status", "CE Remarks", "Drawing Reference",
    "Reference Designators", "Item #", "Part Type", "LTB Date", "Common (AML)",
    "Change Remarks", "Inventory Stock",
]

import re as _re

def _sheet_headers(ws):
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(c).strip() if c is not None else "" for c in row]

@router.post("/verify-template")
async def verify_bom_template(file: UploadFile = File(...)):
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        return {
            "ok": False,
            "message": "1 issue(s) found",
            "issues": {"file": ["Unsupported file type — must be .xlsx or .xls"]},
        }

    content = await file.read()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return {
            "ok": False,
            "message": "1 issue(s) found",
            "issues": {"file": [f"Cannot open workbook: {e}"]},
        }

    issues: dict = {}

    # --- Sheet check ---
    sheet_map = {s.lower(): s for s in wb.sheetnames}
    required_sheets = ["SCRUB BOM", "Settings", "Batch"]
    missing_sheets = [s for s in required_sheets if s.lower() not in sheet_map]
    if missing_sheets:
        issues["structure"] = [f"Missing sheet: {s}" for s in missing_sheets]

    # --- SCRUB BOM headers ---
    scrub_key = sheet_map.get("scrub bom")
    if scrub_key:
        ws = wb[scrub_key]
        found = _sheet_headers(ws)
        found_lower = [h.lower() for h in found]
        missing = [r for r in SCRUB_BOM_REQUIRED if r.lower() not in found_lower]
        hdr_issues = []
        # Duplicates
        seen, dups = set(), set()
        for h in found:
            if h and h in seen:
                dups.add(h)
            seen.add(h)
        for d in dups:
            hdr_issues.append(f"Duplicate header: '{d}'")
        # Missing with typo suggestion
        for m in missing:
            close = difflib.get_close_matches(m.lower(), found_lower, n=1, cutoff=0.7)
            if close:
                actual = found[found_lower.index(close[0])]
                hdr_issues.append(f"Missing header: '{m}' — did you mean '{actual}'?")
            else:
                hdr_issues.append(f"Missing header: '{m}'")
        if hdr_issues:
            issues.setdefault("headers", {})["SCRUB BOM"] = hdr_issues

    # --- Settings headers ---
    settings_key = sheet_map.get("settings")
    if settings_key:
        ws = wb[settings_key]
        found = _sheet_headers(ws)
        found_lower = [h.lower() for h in found]
        hdr_issues = []
        if "fg" not in found_lower:
            hdr_issues.append("Missing header: 'FG'")
        vol_qty_cols = [h for h in found if _re.match(r"V\d+\s*Qty", h, _re.I)]
        if len(vol_qty_cols) < 2:
            hdr_issues.append("Missing at least two volume qty columns (e.g. V1 Qty, V2 Qty)")
        seen, dups = set(), set()
        for h in found:
            if h and h in seen:
                dups.add(h)
            seen.add(h)
        for d in dups:
            hdr_issues.append(f"Duplicate header: '{d}'")
        if hdr_issues:
            issues.setdefault("headers", {})["Settings"] = hdr_issues

    # --- Batch headers ---
    batch_key = sheet_map.get("batch")
    if batch_key:
        ws = wb[batch_key]
        found = _sheet_headers(ws)
        found_lower = [h.lower() for h in found]
        hdr_issues = []
        if "fg" not in found_lower:
            hdr_issues.append("Missing header: 'FG'")
        batch_qty_cols = [h for h in found if _re.match(r"V\d+\s*Batch\s*Qty", h, _re.I)]
        if not batch_qty_cols:
            hdr_issues.append("Missing at least one batch qty column (e.g. V1 Batch Qty)")
        seen, dups = set(), set()
        for h in found:
            if h and h in seen:
                dups.add(h)
            seen.add(h)
        for d in dups:
            hdr_issues.append(f"Duplicate header: '{d}'")
        if hdr_issues:
            issues.setdefault("headers", {})["Batch"] = hdr_issues

    wb.close()

    if not issues:
        return {"ok": True, "message": "Template is valid", "issues": {}}

    total = sum(
        len(v) if isinstance(v, list) else sum(len(x) for x in v.values())
        for v in issues.values()
    )
    return {"ok": False, "message": f"{total} issue(s) found", "issues": issues}


@router.post("/parse-preview")
async def parse_bom_preview(file: UploadFile = File(...)):
    """Quick parse: returns headers + rows for S1 UI preview (no DB save)."""
    ext = file.filename.split('.')[-1].lower()
    content = await file.read()
    if ext in ('csv', 'tsv'):
        import csv as csvmod
        sep = '\t' if ext == 'tsv' else ','
        reader = csvmod.reader(io.StringIO(content.decode('utf-8-sig')), delimiter=sep)
        rows = list(reader)
        if not rows:
            raise HTTPException(400, "Empty file")
        headers = [h.strip() for h in rows[0]]
        data_rows = [r for r in rows[1:] if any(c.strip() for c in r)]
        return {"headers": headers, "dataRows": data_rows, "sheetName": "CSV", "totalRows": len(data_rows)}
    elif ext in ('xlsx', 'xls'):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        for name in wb.sheetnames:
            if "SCRUB" in name.upper() or "BOM" in name.upper():
                ws = wb[name]
                break
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            raise HTTPException(400, "Empty file")
        headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
        data_rows = [[str(c).strip() if c is not None else '' for c in r] for r in all_rows[1:] if any(c is not None for c in r)]
        return {"headers": headers, "dataRows": data_rows, "sheetName": ws.title, "totalRows": len(data_rows)}
    raise HTTPException(400, "Unsupported format")

@router.post("/{project_id}/upload")
async def upload_bom(project_id: int, file: UploadFile = File(...),
                     db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Project).where(Project.id == project_id))
    proj = result.scalar_one_or_none()
    if not proj:
        raise HTTPException(404, detail="Project not found")

    content = await file.read()
    try:
        parsed = parse_scrub_bom(content)
    except Exception as e:
        raise HTTPException(400, detail=f"Failed to parse BOM file: {e}")

    try:
        await db.execute(delete(BomLine).where(BomLine.project_id == project_id))
        lines = []
        for row in parsed["rows"]:
            lines.append(BomLine(project_id=project_id, **row))
        db.add_all(lines)

        # Upsert BomDetail
        await db.execute(delete(BomDetail).where(BomDetail.project_id == project_id))
        bd = parsed["bom_detail"]
        fg_volumes = parsed.get("fg_volumes", {})
        db.add(BomDetail(
            project_id=project_id,
            proto=bd["proto"],
            volume_count=bd["volume_count"],
            total_fg_count=bd["total_fg_count"],
            total_cpn_count=bd["total_cpn_count"],
            is_valid_bom=bd["total_cpn_count"] > 0,
            settings_json=json.dumps(fg_volumes),
        ))

        # Insert FgCount rows
        await db.execute(delete(FgCount).where(FgCount.project_id == project_id))
        fg_count_rows = []
        for fg, vol_map in fg_volumes.items():
            for vol_no, count in vol_map.items():
                fg_count_rows.append(FgCount(
                    project_id=project_id,
                    fg=fg,
                    volume_no=vol_no,
                    count=count,
                ))
        if fg_count_rows:
            db.add_all(fg_count_rows)

        # Update proto_qty on project if Settings sheet provided a proto value
        if bd["proto"] > 0:
            proj.proto_qty = bd["proto"]

        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(400, detail=f"Failed to save BOM lines: {e}")

    return {
        "msg":        "BOM uploaded",
        "total":      parsed["total"],
        "fg_parts":   parsed["fg_parts"],
        "assemblies": parsed["assemblies"],
        "bom_detail": parsed["bom_detail"],
    }

@router.get("/{project_id}/lines")
async def get_bom_lines(project_id: int, db: AsyncSession = Depends(get_db),
                        ):
    result = await db.execute(
        select(BomLine).where(BomLine.project_id == project_id)
                       .order_by(BomLine.fg_part, BomLine.assembly, BomLine.cpn)
    )
    lines = result.scalars().all()
    return [l.__dict__ for l in lines]


@router.delete("/{project_id}/lines")
async def delete_bom_lines(project_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Project).where(Project.id == project_id))
    if not result.scalar_one_or_none():
        raise HTTPException(404, detail="Project not found")
    await db.execute(delete(BomLine).where(BomLine.project_id == project_id))
    await db.commit()
    return {"msg": "BOM lines cleared"}
