"""
QuoteWin / Supplier Quoting Tool — Export File Parser
Also known as: QW, Quoting Tool, Supplier Quoting Tool
Contains: Pricing & lead time from suppliers responding to a Quote Request

File Structure:
  Row 1–9:   Metadata (project info, volume quantities)
             Col A = Project Headings | Col B-C = Empty | Col D = Project Details
  Row 10:    Always empty
  Row 11:    Heading — "Assembly Summary"
  Row 12:    Volume Headings (Cols A–E)
  Row 13:    Board quantities per volume tier  ← parser reads col 2, 6, 10
  Row 14:    Volume Total
  Row 15:    Always empty
  Row 16:    Heading — "Group Detail"
  Row 17:    Supplier Quote Column Headers  ← dynamic regex detection + hardcoded fallback
  Row 18+:   Supplier Quote Data            ← one row per supplier offer per component

Column layout (0-indexed, from row 17 headers):
  0=Project, 1=Part Number, 2=Part Description, 3=Commodity,
  4=Mfg Name, 5=Mfg Part Number,
  6=Cost#1(Conv.), 7=Price(Original)#1,
  8=Cost#2(Conv.), 9=Price(Original)#2,
  10=Cost#3(Conv.), 11=Price(Original)#3,
  12=Currency(Original), 13=Supp Name, 14=Pkg Qty, 15=MOQ,
  16=Lead Time,
  17=Awarded Volume#1, 18=Award#1,
  19=Awarded Volume#2, 20=Award#2,
  21=Awarded Volume#3, 22=Award#3,
  23=Part Qty, 24=Corrected MPN, 25=Long Comment,
  32=NCNR, 37=Part Status
"""
import openpyxl
from collections import defaultdict
from typing import Dict, Any, List, Optional
import io

DATA_START_ROW = 18

def _safe_float(v):
    if v is None: return None
    try:    return float(v)
    except: return None

def _safe_int(v):
    if v is None: return None
    try:    return int(float(v))
    except: return None

def _parse_date(v) -> Optional[str]:
    """Return YYYY-MM-DD string or None."""
    if v is None: return None
    try:
        from datetime import datetime, date
        if isinstance(v, (datetime, date)):
            return v.strftime("%Y-%m-%d")
        s = str(v).strip()
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except: pass
        return s or None
    except: return None

def _price_control(v) -> str:
    """Legacy rule: empty or 'NO CONTRACT' → 'centum', else 'cnp'."""
    if not v: return "centum"
    return "centum" if str(v).strip().upper() in ("", "NO CONTRACT") else "cnp"

def _validity_weeks(eff: Optional[str], exp: Optional[str]) -> Optional[int]:
    """Weeks between two YYYY-MM-DD strings."""
    if not eff or not exp: return None
    try:
        from datetime import datetime
        d1 = datetime.strptime(eff, "%Y-%m-%d")
        d2 = datetime.strptime(exp, "%Y-%m-%d")
        return max(0, (d2 - d1).days // 7)
    except: return None

def _resolve_vol(rows: List[Dict], cost_key: str, award_key: str) -> Dict:
    """
    Per-volume-break price selection:
    1. Award = 100 AND cost > 0  → Awarded
    2. Award = 100 but cost null/0 → fallback to lowest non-zero
    3. No Award = 100 row        → lowest non-zero
    4. All zero/null             → Not Quoted
    """
    awarded_row = next((r for r in rows if r.get(award_key) == 100), None)

    if awarded_row and awarded_row.get(cost_key) and awarded_row[cost_key] > 0:
        return {
            "cost": awarded_row[cost_key],
            "supp": awarded_row["supp"],
            "mpn":  awarded_row["mpn"],
            "moq":  awarded_row["moq"],
            "lt":   awarded_row["lt"],
            "ncnr": awarded_row["ncnr"],
            "currency": awarded_row["currency"],
            "payment_term": awarded_row.get("payment_term", ""),
            "long_comment": awarded_row.get("long_comment", ""),
            "price_control": awarded_row.get("price_control", "centum"),
            "status": "Awarded",
            "note": "",
        }

    quoted = [(r[cost_key], r) for r in rows
              if r.get(cost_key) and r[cost_key] > 0]
    if quoted:
        best_cost, best_row = min(quoted, key=lambda x: x[0])
        if awarded_row:
            note   = f'Award=100 ({awarded_row["supp"]}) had no price — lowest selected'
            status = "Lowest (Award had no price)"
        else:
            note   = "Award=100 not marked for this CPN — lowest price selected"
            status = "Lowest (100 not marked)"
        return {
            "cost": best_cost,
            "supp": best_row["supp"],
            "mpn":  best_row["mpn"],
            "moq":  best_row["moq"],
            "lt":   best_row["lt"],
            "ncnr": best_row["ncnr"],
            "currency": best_row["currency"],
            "payment_term": best_row.get("payment_term", ""),
            "long_comment": best_row.get("long_comment", ""),
            "price_control": best_row.get("price_control", "centum"),
            "status": status,
            "note": note,
        }

    return {
        "cost": None, "supp": "", "mpn": "", "moq": None,
        "lt": None, "ncnr": "", "currency": "", "payment_term": "",
        "long_comment": "", "price_control": "centum",
        "status": "Not Quoted", "note": "",
    }

import re as _re


def parse_qw_file(file_bytes: bytes, has_proto: bool = False) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Read volume break board-level quantities from summary rows (legacy fallback)
    vol1_qty = _safe_int(ws.cell(row=13, column=2).value)
    vol2_qty = _safe_int(ws.cell(row=13, column=6).value)
    vol3_qty = _safe_int(ws.cell(row=13, column=10).value)

    # Build header name→col-index map from row 17
    header_row = list(ws.iter_rows(min_row=17, max_row=17, values_only=True))[0]
    header_map: Dict[str, int] = {}
    for idx, cell_val in enumerate(header_row):
        if cell_val is not None:
            header_map[str(cell_val).strip().lower()] = idx

    # Resolve positional field columns from header_map (with hardcoded fallbacks)
    col_currency      = header_map.get("currency (original)", header_map.get("currency", 12))
    col_supp          = header_map.get("supp name", header_map.get("supplier name", 13))
    col_pkg_qty       = header_map.get("pkg qty", header_map.get("package qty", 14))
    col_moq           = header_map.get("moq", 15)
    col_lt            = header_map.get("lead time", 16)
    col_ncnr          = header_map.get("ncnr", 32)
    col_part_status   = header_map.get("part status", 37)
    col_long_comment  = header_map.get("long comment", 25)
    col_corrected_mpn = header_map.get("corrected mpn", 24)
    col_payment_term  = header_map.get("payment term", header_map.get("payment terms", None))

    # Dynamic column detection for date, price-type, no_bid, cost, price, award, av columns
    eff_col: Optional[int] = None
    exp_col: Optional[int] = None
    price_type_col: Optional[int] = None
    no_bid_col: Optional[int] = None

    # Indexed columns: group_no → col_index
    cost_cols: Dict[int, int] = {}    # Cost #N (Conv.) → group N
    price_cols: Dict[int, int] = {}   # Price (Original) #N → group N
    award_cols: Dict[int, int] = {}   # Award #N → group N
    av_cols: Dict[int, int] = {}      # Awarded Volume #N → group N

    for name, idx in header_map.items():
        if "eff" in name and "date" in name and eff_col is None:
            eff_col = idx
        elif "exp" in name and "date" in name and exp_col is None:
            exp_col = idx
        # fallback: any eff/exp without "date"
        if eff_col is None and "eff" in name:
            eff_col = idx
        if exp_col is None and "exp" in name:
            exp_col = idx
        if "price type" in name and price_type_col is None:
            price_type_col = idx
        if ("no bid" in name or "nobid" in name) and no_bid_col is None:
            no_bid_col = idx

        m = _re.search(r"cost\s*#\s*(\d+)", name)
        if m:
            cost_cols[int(m.group(1))] = idx
        m = _re.search(r"price\s*\(original\)\s*#\s*(\d+)", name)
        if m:
            price_cols[int(m.group(1))] = idx
        m = _re.fullmatch(r"award\s*#\s*(\d+)", name)
        if m:
            award_cols[int(m.group(1))] = idx
        m = _re.search(r"awarded\s*volume\s*#\s*(\d+)", name)
        if m:
            av_cols[int(m.group(1))] = idx

    # Fallback to positional columns if dynamic detection found nothing
    if not cost_cols:
        cost_cols  = {1: 6, 2: 8,  3: 10}
        price_cols = {1: 7, 2: 9,  3: 11}
        award_cols = {1: 18, 2: 20, 3: 22}
        av_cols    = {1: 17, 2: 19, 3: 21}

    found_vol_count = len(award_cols)

    # Build volume_map: Award group index → logical volume_no (1, 2, 3)
    # has_proto=True: group 1 = proto (skip for production), group 2→vol1, group 3→vol2, etc.
    # has_proto=False: group 1→vol1, group 2→vol2, group 3→vol3
    sorted_groups = sorted(award_cols.keys())
    if has_proto and sorted_groups:
        proto_group = sorted_groups[0]
        volume_map = {g: i for i, g in enumerate(sorted_groups[1:], start=1)}
    else:
        proto_group = None
        volume_map = {g: i for i, g in enumerate(sorted_groups, start=1)}

    def _cell(row, col_idx):
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx]

    # Group rows by CPN
    parts: Dict[str, List[Dict]] = defaultdict(list)
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        cpn = row[1]
        if not cpn:
            continue
        cpn = str(cpn).strip()

        eff_date = _parse_date(_cell(row, eff_col)) if eff_col is not None else None
        exp_date = _parse_date(_cell(row, exp_col)) if exp_col is not None else None
        pc_raw   = _cell(row, price_type_col)

        # Build cost/award data per group
        group_data: Dict[int, Dict] = {}
        for g in sorted_groups:
            group_data[g] = {
                "cost":      _safe_float(_cell(row, cost_cols.get(g))),
                "price_orig": _safe_float(_cell(row, price_cols.get(g))),
                "award":     _safe_int(_cell(row, award_cols.get(g))),
                "av":        _safe_int(_cell(row, av_cols.get(g))),
            }

        # Proto pricing (from proto_group)
        proto_cost_conv = None
        proto_price_orig_val = None
        if has_proto and proto_group is not None:
            proto_cost_conv   = group_data[proto_group]["cost"]
            proto_price_orig_val = group_data[proto_group]["price_orig"]

        # Map groups to vol keys using volume_map
        vol_data: Dict[int, Dict] = {vol_no: group_data[g] for g, vol_no in volume_map.items()}

        record = {
            "mpn":              str(row[5] or "").strip(),
            "currency":         str(_cell(row, col_currency) or "USD").strip(),
            "supp":             str(_cell(row, col_supp) or "").strip(),
            "pkg_qty":          _safe_int(_cell(row, col_pkg_qty)),
            "moq":              _safe_int(_cell(row, col_moq)),
            "lt":               _safe_int(_cell(row, col_lt)),
            "ncnr":             str(_cell(row, col_ncnr) or "").strip(),
            "part_status":      str(_cell(row, col_part_status) or "").strip(),
            "payment_term":     str(_cell(row, col_payment_term) or "").strip() if col_payment_term is not None else "",
            "long_comment":     str(_cell(row, col_long_comment) or "").strip(),
            "corrected_mpn":    str(_cell(row, col_corrected_mpn) or "").strip(),
            "part_description": str(row[2] or "").strip() if len(row) > 2 else "",
            "eff_date":         eff_date,
            "exp_date":         exp_date,
            "price_control":    _price_control(pc_raw),
            "no_bid":           bool(_cell(row, no_bid_col)) if no_bid_col is not None else False,
            # Production volume cost/award using mapped vol_no keys
            "cost1":            vol_data.get(1, {}).get("cost"),
            "cost2":            vol_data.get(2, {}).get("cost"),
            "cost3":            vol_data.get(3, {}).get("cost"),
            "price1_orig":      vol_data.get(1, {}).get("price_orig"),
            "price2_orig":      vol_data.get(2, {}).get("price_orig"),
            "price3_orig":      vol_data.get(3, {}).get("price_orig"),
            "av1":              vol_data.get(1, {}).get("av"),
            "award1":           vol_data.get(1, {}).get("award"),
            "av2":              vol_data.get(2, {}).get("av"),
            "award2":           vol_data.get(2, {}).get("award"),
            "av3":              vol_data.get(3, {}).get("av"),
            "award3":           vol_data.get(3, {}).get("award"),
            "proto_cost_conv":  proto_cost_conv,
            "proto_price_orig": proto_price_orig_val,
        }
        parts[cpn].append(record)

    # Resolve each CPN per volume break
    resolved: Dict[str, Dict] = {}
    for cpn, rows in parts.items():
        awarded_supp = next(
            (r["supp"] for r in rows if r["award1"] == 100), ""
        )
        v1 = _resolve_vol(rows, "cost1", "award1")
        v2 = _resolve_vol(rows, "cost2", "award2")
        v3 = _resolve_vol(rows, "cost3", "award3")
        resolved[cpn] = {
            "cpn":          cpn,
            "awarded_supp": awarded_supp,
            "v1": v1, "v2": v2, "v3": v3,
        }

    # Build flat list for DB insert
    db_rows: List[Dict] = []
    for cpn, rows in parts.items():
        for r in rows:
            eff = r.get("eff_date")
            exp = r.get("exp_date")
            db_rows.append({
                "cpn":                  cpn,
                "mpn":                  r["mpn"],
                "supp_name":            r["supp"],
                "currency":             r["currency"],
                "cost1_conv":           r["cost1"],
                "cost2_conv":           r["cost2"],
                "cost3_conv":           r["cost3"],
                "price1_orig":          r["price1_orig"],
                "price2_orig":          r["price2_orig"],
                "price3_orig":          r["price3_orig"],
                "moq":                  r["moq"],
                "pkg_qty":              r["pkg_qty"],
                "lead_time":            r["lt"],
                "award1":               r["award1"],
                "award2":               r["award2"],
                "award3":               r["award3"],
                "awarded_vol1":         r["av1"],
                "awarded_vol2":         r["av2"],
                "awarded_vol3":         r["av3"],
                "ncnr":                 r["ncnr"],
                "part_status":          r["part_status"],
                "payment_term":         r["payment_term"],
                "long_comment":         r["long_comment"],
                "corrected_mpn":        r["corrected_mpn"],
                "part_description":     r["part_description"],
                "effective_from_date":  eff,
                "expiry_date":          exp,
                "quote_validity_weeks": _validity_weeks(eff, exp),
                "price_control":        r["price_control"],
                "no_bid":               r["no_bid"],
                "proto_cost_conv":      r.get("proto_cost_conv"),
                "proto_price_orig":     r.get("proto_price_orig"),
            })

    # When has_proto: vol1_qty from QW row 13 corresponds to prod vol1 (not proto)
    # vol quantities from QW summary rows remain available for caller to use as needed
    return {
        "vol1_qty":        vol1_qty,
        "vol2_qty":        vol2_qty,
        "vol3_qty":        vol3_qty,
        "resolved":        resolved,
        "db_rows":         db_rows,
        "cpn_count":       len(resolved),
        "found_vol_count": found_vol_count,
        "has_proto":       has_proto,
        "cpns_in_file":    set(resolved.keys()),
    }
