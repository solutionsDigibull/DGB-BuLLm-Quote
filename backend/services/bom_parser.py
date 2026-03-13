"""
Parses the Knorr Bremse SCRUB BOM Excel file.
Column mapping (0-indexed):
  A=FG, B=Level, C=Assembly, D=CPN, E=Description, F=MFR, G=MPN,
  H=Commodity, I=Qty, J=UOM, K=Part Status, L=CE Remarks,
  M=Drawing Reference, N=Ref Des, O=Item#, P=Part Type,
  Q=LTB Date, R=Common with MPN, S=Change Remarks,
  T=Inventory Stock, U=Common(AML)
"""
import openpyxl
from typing import List, Dict, Any, Optional
import io
import re

HEADER_ROW = 1   # Row index (1-based) where headers live in SCRUB BOM sheet

COL_MAP = {
    "fg":          0,   # A
    "level":       1,   # B
    "assembly":    2,   # C
    "cpn":         3,   # D
    "description": 4,   # E
    "manufacturer":5,   # F
    "mpn":         6,   # G
    "commodity":   7,   # H
    "qty":         8,   # I
    "uom":         9,   # J
    "part_status": 10,  # K
    "ltb_date":    16,  # Q
    "ref_des":     13,  # N
    "inventory_stock": 19,  # T
}

HEADER_PATTERNS = {
    "fg":           [r"^fg$", r"finished.?good", r"top.?level", r"^fg.?part"],
    "assembly":     [r"^assembly$", r"^assy$", r"sub.?assy", r"parent.?assy"],
    "level":        [r"^level$", r"^lvl$", r"assembly.?level"],
    "cpn":          [r"^cpn$", r"item.?code", r"cust.?p"],
    "description":  [r"^desc", r"description"],
    "manufacturer": [r"^mfr$", r"^mfg$", r"manufacturer"],
    "mpn":          [r"^mpn$", r"part.?num", r"mfr.?pn"],
    "commodity":    [r"^commodity$", r"^category$"],
    "qty":          [r"^qty", r"^quantity", r"^qpa$", r"bom.?qty"],
    "uom":          [r"^uom$", r"unit.?of.?meas"],
    "part_status":  [r"part.?status", r"lifecycle", r"^status$"],
    "ltb_date":     [r"ltb", r"last.?time.?buy"],
    "ref_des":      [r"ref.?des", r"designator"],
    "inventory_stock": [r"inv", r"stock", r"on.?hand"],
}


def _detect_columns(header_row) -> Optional[Dict[str, int]]:
    """Detect column indices from header names. Returns dict or None."""
    headers = [str(h).strip().lower() if h else "" for h in header_row]
    detected = {}
    used_cols = set()
    for field, patterns in HEADER_PATTERNS.items():
        for i, h in enumerate(headers):
            if i in used_cols:
                continue
            for pat in patterns:
                if re.search(pat, h, re.IGNORECASE):
                    detected[field] = i
                    used_cols.add(i)
                    break
            if field in detected:
                break
    return detected if len(detected) >= 3 else None


def _parse_settings_sheet(wb) -> Dict[str, Any]:
    """
    Read the Settings sheet for proto qty and per-FG volume quantities.
    Expected layout:
      - First column: row labels (e.g. "Proto", "FG1", "FG2", ...)
      - Header row contains "FG" and "V1 Qty", "V2 Qty", ... columns
    Returns:
      {
        "proto": int,
        "volume_count": int,
        "fg_volumes": {fg: {1: qty, 2: qty, ...}}
      }
    """
    default = {"proto": 0, "volume_count": 2, "fg_volumes": {}}

    settings_ws = None
    for name in wb.sheetnames:
        if "setting" in name.lower():
            settings_ws = wb[name]
            break
    if not settings_ws:
        return default

    all_rows = list(settings_ws.iter_rows(values_only=True))
    if not all_rows:
        return default

    # Find header row: look for a row containing volume qty headers (V\d+ Qty)
    header_idx = None
    header_row = None
    vol_col_map: Dict[int, int] = {}   # volume_no → col_index
    fg_col: Optional[int] = None
    proto_col: Optional[int] = None

    for ri, row in enumerate(all_rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        # Check for "V1 Qty" style headers
        vol_matches = [(ci, re.match(r"V(\d+)\s*Qty", c, re.IGNORECASE)) for ci, c in enumerate(cells)]
        vol_matches = [(ci, m) for ci, m in vol_matches if m]
        if vol_matches:
            header_idx = ri
            header_row = cells
            for ci, m in vol_matches:
                vol_col_map[int(m.group(1))] = ci
            # Find FG column
            for ci, c in enumerate(cells):
                if re.search(r"^fg$", c, re.IGNORECASE):
                    fg_col = ci
                    break
            # Find proto column
            for ci, c in enumerate(cells):
                if re.search(r"proto", c, re.IGNORECASE):
                    proto_col = ci
                    break
            break

    if header_idx is None or not vol_col_map:
        return default

    volume_count = len(vol_col_map)
    proto_qty = 0
    fg_volumes: Dict[str, Dict[int, float]] = {}

    for row in all_rows[header_idx + 1:]:
        cells = [row[i] if i < len(row) else None for i in range(len(all_rows[header_idx]))]

        # Check if this is the proto row
        label = str(cells[0]).strip().upper() if cells[0] else ""
        if label == "PROTO" or (proto_col is not None and cells[proto_col] is not None):
            # Proto row: read first vol qty column as proto qty
            if vol_col_map:
                first_vol_ci = vol_col_map.get(1, min(vol_col_map.values()))
                v = cells[first_vol_ci] if first_vol_ci < len(cells) else None
                try:
                    proto_qty = int(float(v)) if v is not None else 0
                except (ValueError, TypeError):
                    proto_qty = 0
            if label == "PROTO":
                continue  # don't add proto as an FG

        # Read FG volumes
        fg_val = None
        if fg_col is not None and fg_col < len(cells):
            fg_val = cells[fg_col]
        elif cells[0]:
            fg_val = cells[0]

        if not fg_val:
            continue
        fg_str = str(fg_val).strip()
        if not fg_str or fg_str.upper() in ("FG", "PROTO", ""):
            continue

        vols: Dict[int, float] = {}
        for vol_no, ci in vol_col_map.items():
            v = cells[ci] if ci < len(cells) else None
            try:
                vols[vol_no] = float(v) if v is not None else 0.0
            except (ValueError, TypeError):
                vols[vol_no] = 0.0
        if any(v > 0 for v in vols.values()):
            fg_volumes[fg_str] = vols

    return {
        "proto": proto_qty,
        "volume_count": volume_count,
        "fg_volumes": fg_volumes,
    }


def parse_scrub_bom(file_bytes: bytes) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Sheet selection: prefer 'SCRUB BOM', else first sheet
    target_sheet = None
    for name in wb.sheetnames:
        if "SCRUB" in name.upper() or "BOM" in name.upper():
            target_sheet = wb[name]
            break
    if not target_sheet:
        target_sheet = wb.active

    # Try header-based detection, fallback to positional COL_MAP
    header_row = next(target_sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    col_map = (_detect_columns(header_row) if header_row else None) or COL_MAP

    rows_out: List[Dict] = []
    fg_parts = set()
    assemblies = set()

    for row in target_sheet.iter_rows(min_row=2, values_only=True):
        cpn = row[col_map["cpn"]] if "cpn" in col_map and col_map["cpn"] < len(row) else None
        if not cpn:
            continue

        def _cell(field, default=""):
            i = col_map.get(field)
            if i is None or i >= len(row):
                return default
            return row[i]

        cpn = str(cpn).strip()
        fg  = str(_cell("fg") or "").strip()
        asm = str(_cell("assembly") or cpn).strip()

        # Determine level
        level_raw = _cell("level")
        try:
            level = int(level_raw) if level_raw is not None else 1
        except (ValueError, TypeError):
            level = 1

        qty_raw = _cell("qty")
        try:
            qty = float(qty_raw) if qty_raw is not None else 0.0
        except (ValueError, TypeError):
            qty = 0.0

        inv_raw = _cell("inventory_stock")
        try:
            inv = float(inv_raw) if inv_raw is not None else 0.0
        except (ValueError, TypeError):
            inv = 0.0

        ltb_raw = _cell("ltb_date")
        ltb = str(ltb_raw).strip() if ltb_raw else ""

        record = {
            "fg_part":        fg,
            "assembly":       asm,
            "cpn":            cpn,
            "description":    str(_cell("description") or "").strip(),
            "manufacturer":   str(_cell("manufacturer") or "").strip(),
            "mpn":            str(_cell("mpn") or "").strip(),
            "commodity":      str(_cell("commodity") or "").strip(),
            "qty":            qty,
            "uom":            str(_cell("uom") or "NUM").strip(),
            "part_status":    str(_cell("part_status") or "").strip(),
            "ltb_date":       ltb,
            "ref_des":        str(_cell("ref_des") or "").strip(),
            "inventory_stock":inv,
            "level":          level,
        }
        rows_out.append(record)
        if fg:
            fg_parts.add(fg)
        if asm:
            assemblies.add(asm)

    settings = _parse_settings_sheet(wb)

    return {
        "rows":       rows_out,
        "fg_parts":   sorted(fg_parts),
        "assemblies": sorted(assemblies),
        "total":      len(rows_out),
        "bom_detail": {
            "proto":           settings["proto"],
            "volume_count":    settings["volume_count"],
            "total_fg_count":  len(fg_parts),
            "total_cpn_count": len(rows_out),
        },
        "fg_volumes": settings["fg_volumes"],
    }
