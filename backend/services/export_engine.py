"""
BuLLMQuote Excel Export Engine
Generates a single workbook with 15 sheets matching the Centum Internal CBOM format.

Sheets:
  1.  CBOM Proto
  2.  CBOM VL-1
  3.  CBOM VL-2
  4.  BOM MATRIX - Proto
  5.  BOM MATRIX - VL1
  6.  BOM MATRIX - VL2
  7.  Part Num vs Mfg Num
  8.  NRE Charges
  9.  Missing Notes
  10. Lead Time FG Wise
  11. Sum & Count
  12. Sum & Count Detailed
  13. Revision History
  14. Price Control
  15. Summary
  16. Batchwise Cashflow
  17. Excess Inventory Proto
  18. Excess Inventory VL-1
  19. Excess Inventory VL-2
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from typing import Dict, List, Any, Optional
from datetime import datetime
import math, io

# ── Style constants ─────────────────────────────────────────────────────────
FONT = "Arial"
HDR_DARK   = "1A3A5C"
HDR_MED    = "2C5282"
HDR_LIGHT  = "BDD7EE"
ROW_ALT    = "F2F7FC"
ROW_WHITE  = "FFFFFF"
GREEN_BG   = "E8F5E9"; GREEN_FG   = "1B5E20"
YELLOW_BG  = "FFF9C4"; YELLOW_FG  = "7C5800"
ORANGE_BG  = "FFF3E0"; ORANGE_FG  = "8D4E00"
RED_BG     = "FFEBEE"; RED_FG     = "C62828"
STATUS_COLORS = {
    "Awarded":                    (GREEN_BG,  GREEN_FG),
    "Lowest (Award had no price)":(YELLOW_BG, YELLOW_FG),
    "Lowest (100 not marked)":    (ORANGE_BG, ORANGE_FG),
    "Not Quoted":                 (RED_BG,    RED_FG),
}

def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_c):
    return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def _font(bold=False, size=9, color="000000", italic=False):
    return Font(name=FONT, bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr_cell(ws, row, col, value, bg=HDR_DARK, fg="FFFFFF", bold=True,
              align="center", size=9, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, size=size, color=fg)
    c.fill      = _fill(bg)
    c.alignment = _align(h=align, wrap=wrap)
    c.border    = _thin()
    return c

def _data_cell(ws, row, col, value, bg=ROW_WHITE, fg="000000", bold=False,
               align="left", num_fmt=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, color=fg)
    c.fill      = _fill(bg)
    c.alignment = _align(h=align, wrap=wrap)
    c.border    = _thin()
    if num_fmt:
        c.number_format = num_fmt
    return c

NUM_PRICE = '#,##0.00000;[Red](#,##0.00000);"-"'
NUM_QTY   = '#,##0;"-"'
NUM_PCT   = '0.00%;"-"'

# ── CBOM sheet builder ───────────────────────────────────────────────────────
CBOM_HEADERS = [
    ("FG Part No.",      14),
    ("Assembly",         18),
    ("CPN",              20),
    ("Description",      36),
    ("Commodity",        14),
    ("Mfg Name",         24),
    ("Mfg Part Number",  28),
    ("UOM",               6),
    ("Part Qty",          9),
    ("Ext Vol Qty",      10),
    ("Unit Price\n(USD Conv.)", 14),
    ("Price Orig.",      12),
    ("Currency",          9),
    ("Ext Price\n(Conv.)",12),
    ("Ext Vol Price\n(Conv.)",13),
    ("Supplier",         28),
    ("Pkg Qty",           8),
    ("MOQ",               8),
    ("Lead Time\n(wks)",  9),
    ("Stock",             8),
    ("NRE Charge\n(Conv.)",11),
    ("NCNR",              7),
    ("Price Status",     26),
    ("Scrap\nFactor",     9),
    ("Payment Term",     18),
    ("Notes",            32),
]

def _build_cbom_sheet(wb, sheet_name, cbom_rows, board_qty, project, assembly_summary_data):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A20"

    # ── Header block ──
    ws.merge_cells("A1:Z1")
    c = ws.cell(row=1, column=1, value="Awards Costed BOM")
    c.font = _font(bold=True, size=12, color="FFFFFF")
    c.fill = _fill(HDR_DARK)
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 22

    meta = [
        ("Template:", "Centum Costed BOM_Internal"),
        ("Project:",  f'{project["code"]} - {project["customer"]}'),
        ("Volume:",   str(board_qty)),
        ("Currency:", project.get("currency", "USD")),
        ("Report On:","Selected Cost"),
        ("Sort By:",  "Part Number"),
    ]
    for i, (lbl, val) in enumerate(meta, 2):
        ws.cell(row=i, column=3, value=lbl).font = _font(bold=True, size=9)
        ws.cell(row=i, column=6, value=val).font  = _font(size=9)
    ws.cell(row=2, column=10, value="Currency")
    ws.cell(row=2, column=11, value="Value")
    ws.cell(row=3, column=10, value="EUR"); ws.cell(row=3, column=11, value=project.get("eur_rate", 0.8585))
    ws.cell(row=4, column=10, value="INR"); ws.cell(row=4, column=11, value=project.get("inr_rate", 89.47))
    ws.cell(row=5, column=10, value="USD"); ws.cell(row=5, column=11, value=1)

    # ── Assembly summary ──
    ws.cell(row=9, column=3, value="*** Assembly Summary").font = _font(bold=True)
    for i, (asm, sd) in enumerate(assembly_summary_data.items(), 10):
        ws.cell(row=i, column=3, value=asm).font = _font(bold=True)
        ws.cell(row=i, column=5, value=round(sd["ext_price_sum"], 6))
        ws.cell(row=i, column=6, value=board_qty)
        ws.cell(row=i, column=7, value=round(sd["ext_price_sum"] * board_qty, 6))

    # ── Column headers row 19 ──
    for col, (hdr, width) in enumerate(CBOM_HEADERS, 1):
        _hdr_cell(ws, 19, col, hdr, wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[19].height = 32
    ws.auto_filter.ref = f"A19:{get_column_letter(len(CBOM_HEADERS))}19"

    # ── Data rows ──
    current_asm = None
    data_row = 20
    for i, row in enumerate(cbom_rows):
        # Assembly separator row
        if row["assembly"] != current_asm:
            current_asm = row["assembly"]
            ws.merge_cells(f"C{data_row}:Z{data_row}")
            sep = ws.cell(row=data_row, column=3,
                          value=f"*** Assembly: {current_asm}")
            sep.font = _font(bold=True, italic=True, size=9, color="FFFFFF")
            sep.fill = _fill(HDR_MED)
            data_row += 1

        bg = ROW_ALT if i % 2 else ROW_WHITE
        sc = STATUS_COLORS.get(row.get("price_status",""), (bg, "000000"))

        def dc(col, val, fmt=None, align="left"):
            _data_cell(ws, data_row, col, val, bg=sc[0], fg=sc[1],
                       align=align, num_fmt=fmt)

        dc(1,  row.get("fg_part",""))
        dc(2,  row.get("assembly",""))
        dc(3,  row.get("cpn",""),      bold=True)
        dc(4,  row.get("description",""), wrap=True)
        dc(5,  row.get("commodity",""))
        dc(6,  row.get("manufacturer",""))
        dc(7,  row.get("mpn",""))
        dc(8,  row.get("uom",""),      align="center")
        dc(9,  row.get("part_qty"),    fmt=NUM_QTY, align="right")
        dc(10, row.get("ext_vol_qty"), fmt=NUM_QTY, align="right")
        dc(11, row.get("unit_price_conv"), fmt=NUM_PRICE, align="right")
        dc(12, row.get("price_orig"),  fmt=NUM_PRICE, align="right")
        dc(13, row.get("currency_orig",""), align="center")
        dc(14, row.get("ext_price_conv"),  fmt=NUM_PRICE, align="right")
        dc(15, row.get("ext_vol_price"),   fmt=NUM_PRICE, align="right")
        dc(16, row.get("supp_name",""))
        dc(17, row.get("pkg_qty"),     fmt=NUM_QTY, align="right")
        dc(18, row.get("moq"),         fmt=NUM_QTY, align="right")
        dc(19, row.get("lead_time"),   fmt=NUM_QTY, align="center")
        dc(20, row.get("stock"),       fmt=NUM_QTY, align="right")
        dc(21, row.get("nre_charge_conv"), fmt=NUM_PRICE, align="right")
        dc(22, row.get("ncnr",""),     align="center")
        # Status cell coloured by its own status
        sc2 = STATUS_COLORS.get(row.get("price_status",""), (bg, "000000"))
        sc_cell = ws.cell(row=data_row, column=23, value=row.get("price_status",""))
        sc_cell.font = _font(bold=True, color=sc2[1])
        sc_cell.fill = _fill(sc2[0])
        sc_cell.alignment = _align()
        sc_cell.border = _thin()
        dc(24, row.get("scrap_factor"), fmt=NUM_PCT, align="center")
        dc(25, row.get("payment_term",""))
        dc(26, row.get("long_comment",""), wrap=True)

        ws.row_dimensions[data_row].height = 15
        data_row += 1

    return ws

# ── BOM MATRIX sheet ─────────────────────────────────────────────────────────
def _build_bom_matrix(wb, sheet_name, cbom_rows, board_qty):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Get unique assemblies (columns) and unique CPNs (rows)
    assemblies = list(dict.fromkeys(r["assembly"] for r in cbom_rows))
    # Group by CPN
    cpn_data: Dict[str, Dict] = {}
    for r in cbom_rows:
        cpn = r["cpn"]
        if cpn not in cpn_data:
            cpn_data[cpn] = {
                "desc":      r["description"],
                "commodity": r["commodity"],
                "mfg":       r["manufacturer"],
                "mpn":       r["mpn"],
                "unit_price":r.get("unit_price_conv", 0) or 0,
                "qty_by_asm":{},
            }
        cpn_data[cpn]["qty_by_asm"][r["assembly"]] = r.get("part_qty", 0) or 0

    # Header
    hdr_cols = ["Part Number", "Grand Total", "EAU", "Description", "Commodity",
                "Mfg Name", "Mfg Part Number", "Unit price"] + assemblies
    for col, h in enumerate(hdr_cols, 1):
        c = _hdr_cell(ws, 5, col, h, wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = (
            20 if col == 1 else 12 if col <= 8 else 18
        )
    ws.row_dimensions[5].height = 30

    # Demand row
    ws.cell(row=3, column=1, value="Demand")
    for ai, asm in enumerate(assemblies):
        ws.cell(row=3, column=9+ai, value=board_qty)

    # Data
    for ri, (cpn, d) in enumerate(cpn_data.items(), 6):
        bg = ROW_ALT if ri % 2 else ROW_WHITE
        grand_total = sum(d["qty_by_asm"].values())
        eau         = grand_total * board_qty

        _data_cell(ws, ri, 1, cpn,           bg=bg, bold=True)
        _data_cell(ws, ri, 2, grand_total,   bg=bg, align="right", num_fmt=NUM_QTY)
        _data_cell(ws, ri, 3, eau,           bg=bg, align="right", num_fmt=NUM_QTY)
        _data_cell(ws, ri, 4, d["desc"],     bg=bg)
        _data_cell(ws, ri, 5, d["commodity"],bg=bg)
        _data_cell(ws, ri, 6, d["mfg"],      bg=bg)
        _data_cell(ws, ri, 7, d["mpn"],      bg=bg)
        _data_cell(ws, ri, 8, d["unit_price"],bg=bg,align="right",num_fmt=NUM_PRICE)
        for ai, asm in enumerate(assemblies):
            qty = d["qty_by_asm"].get(asm, 0)
            ext = qty * d["unit_price"]
            _data_cell(ws, ri, 9+ai, ext, bg=bg, align="right", num_fmt=NUM_PRICE)
        ws.row_dimensions[ri].height = 15

    ws.auto_filter.ref = f"A5:{get_column_letter(len(hdr_cols))}5"
    ws.freeze_panes = "A6"
    return ws

# ── Part Num vs Mfg Num ──────────────────────────────────────────────────────
def _build_part_vs_mfg(wb, bom_lines, qw_resolved):
    ws = wb.create_sheet("Part Num vs Mfg Num")
    ws.sheet_view.showGridLines = False
    for col, (h, w) in enumerate([("Part Number",18),("Mfg Part Number",28),
                                    ("In Scrub BOM",14),("QW MPN",28),("Match",10)], 1):
        _hdr_cell(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    bom_cpns = {l["cpn"]: l["mpn"] for l in bom_lines}
    all_cpns = sorted(set(list(bom_cpns.keys()) + list(qw_resolved.keys())))

    for ri, cpn in enumerate(all_cpns, 2):
        in_bom  = cpn in bom_cpns
        bom_mpn = bom_cpns.get(cpn, "")
        qw_mpn  = qw_resolved.get(cpn, {}).get("v1", {}).get("mpn", "")
        match   = (bom_mpn.strip().upper() == qw_mpn.strip().upper()) if bom_mpn and qw_mpn else False
        bg = ROW_ALT if ri % 2 else ROW_WHITE
        flag_bg = GREEN_BG if match else (YELLOW_BG if in_bom else RED_BG)

        _data_cell(ws, ri, 1, cpn,     bg=bg, bold=True)
        _data_cell(ws, ri, 2, bom_mpn, bg=bg)
        _data_cell(ws, ri, 3, in_bom,  bg=bg, align="center")
        _data_cell(ws, ri, 4, qw_mpn,  bg=bg)
        _data_cell(ws, ri, 5, "✓" if match else "✗", bg=flag_bg,
                   fg=GREEN_FG if match else RED_FG, align="center", bold=True)
        ws.row_dimensions[ri].height = 15
    return ws

# ── NRE Charges ──────────────────────────────────────────────────────────────
def _build_nre(wb, nre_lines):
    ws = wb.create_sheet("NRE Charges")
    ws.sheet_view.showGridLines = False
    headers = [("NRE Type",14),("CPN",20),("Description",36),("Commodity",14),
               ("Mfg Name",24),("Mfg Part Number",28),("NRE Charge (Conv.) USD",18)]
    for col,(h,w) in enumerate(headers,1):
        _hdr_cell(ws,1,col,h)
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.row_dimensions[1].height=20

    total = 0.0
    for ri,n in enumerate(nre_lines,2):
        bg = ROW_ALT if ri%2 else ROW_WHITE
        _data_cell(ws,ri,1,n.get("nre_type",""),bg=bg)
        _data_cell(ws,ri,2,n.get("cpn",""),bg=bg,bold=True)
        _data_cell(ws,ri,3,n.get("description",""),bg=bg)
        _data_cell(ws,ri,4,n.get("commodity",""),bg=bg)
        _data_cell(ws,ri,5,n.get("manufacturer",""),bg=bg)
        _data_cell(ws,ri,6,n.get("mpn",""),bg=bg)
        _data_cell(ws,ri,7,n.get("nre_charge_conv",0),bg=bg,
                   align="right",num_fmt=NUM_PRICE)
        total += n.get("nre_charge_conv",0) or 0
        ws.row_dimensions[ri].height=15

    tr = len(nre_lines)+2
    _hdr_cell(ws,tr,6,"TOTAL",bg=HDR_DARK)
    _hdr_cell(ws,tr,7,total,bg=HDR_DARK,align="right")
    ws.cell(row=tr,column=7).number_format=NUM_PRICE
    return ws

# ── Missing Notes ─────────────────────────────────────────────────────────────
def _build_missing_notes(wb, cbom_rows_proto):
    ws = wb.create_sheet("Missing Notes")
    ws.sheet_view.showGridLines = False
    sections = [
        ("Below PCB Part is NCNR",  lambda r: r.get("ncnr") and str(r["ncnr"]).strip()),
        ("Not Quoted",              lambda r: r.get("price_status") == "Not Quoted"),
        ("Missing MOQ",             lambda r: not r.get("moq")),
        ("Missing Lead Time",       lambda r: not r.get("lead_time")),
    ]
    hdrs = ["Sl.no","CPN","Description","Commodity","Mfg Name","Mfg Part Number",
            "Supplier","Price Status","Remark"]
    row_cursor = 1
    for section_title, predicate in sections:
        matches = [r for r in cbom_rows_proto if predicate(r)]
        if not matches:
            continue
        ws.merge_cells(f"A{row_cursor}:I{row_cursor}")
        t = ws.cell(row=row_cursor,column=1,value=section_title)
        t.font=_font(bold=True,color="FFFFFF"); t.fill=_fill(HDR_MED)
        row_cursor+=1
        for col,(h,w) in enumerate(zip(hdrs,[6,20,36,14,24,28,24,20,20]),1):
            _hdr_cell(ws,row_cursor,col,h)
            ws.column_dimensions[get_column_letter(col)].width=w
        row_cursor+=1
        for si,r in enumerate(matches,1):
            bg=ROW_ALT if si%2 else ROW_WHITE
            _data_cell(ws,row_cursor,1,si,bg=bg,align="center")
            _data_cell(ws,row_cursor,2,r.get("cpn",""),bg=bg,bold=True)
            _data_cell(ws,row_cursor,3,r.get("description",""),bg=bg)
            _data_cell(ws,row_cursor,4,r.get("commodity",""),bg=bg)
            _data_cell(ws,row_cursor,5,r.get("manufacturer",""),bg=bg)
            _data_cell(ws,row_cursor,6,r.get("mpn",""),bg=bg)
            _data_cell(ws,row_cursor,7,r.get("supp_name",""),bg=bg)
            _data_cell(ws,row_cursor,8,r.get("price_status",""),bg=bg)
            _data_cell(ws,row_cursor,9,section_title,bg=bg)
            ws.row_dimensions[row_cursor].height=15
            row_cursor+=1
        row_cursor+=1
    return ws

# ── Lead Time FG Wise ─────────────────────────────────────────────────────────
def _build_lead_time_fg(wb, cbom_rows_proto):
    ws = wb.create_sheet("Lead Time FG Wise")
    ws.sheet_view.showGridLines = False
    fg_parts = list(dict.fromkeys(r["fg_part"] for r in cbom_rows_proto if r.get("fg_part")))
    # Per FG: count of parts by lead time band
    col_offset = 1
    for fg in fg_parts:
        rows_fg = [r for r in cbom_rows_proto if r.get("fg_part") == fg]
        lt_counts: Dict[int,int] = defaultdict(int)
        for r in rows_fg:
            lt = r.get("lead_time") or 0
            lt_counts[int(lt)] += 1
        ws.cell(row=1, column=col_offset,   value="FG part number")
        ws.cell(row=1, column=col_offset+1, value=fg)
        ws.cell(row=3, column=col_offset,   value="LT in weeks")
        ws.cell(row=3, column=col_offset+1, value="Count of Part Number")
        for row_i, (lt_val, count) in enumerate(sorted(lt_counts.items()), 4):
            ws.cell(row=row_i, column=col_offset,   value=lt_val)
            ws.cell(row=row_i, column=col_offset+1, value=count)
        col_offset += 3
    return ws

# ── Sum & Count ───────────────────────────────────────────────────────────────
def _build_sum_count(wb, cbom_proto, cbom_vl1, cbom_vl2, bom_lines):
    ws = wb.create_sheet("Sum & Count")
    ws.sheet_view.showGridLines = False
    hdrs = ["Assembly","Scrub BOM Sum","Proto Sum","VL-1 Sum","VL-2 Sum",
            "Scrub BOM Count","Proto Count","VL-1 Count","VL-2 Count",
            "Is Sum Same","Is Count Same"]
    for col,h in enumerate(hdrs,1):
        _hdr_cell(ws,1,col,h)
        ws.column_dimensions[get_column_letter(col)].width=16
    ws.row_dimensions[1].height=20

    def _asm_sum(rows,asm):
        return sum(r.get("part_qty",0) or 0 for r in rows if r.get("assembly")==asm)
    def _asm_count(rows,asm):
        return sum(1 for r in rows if r.get("assembly")==asm)
    def _bom_sum(lines,asm):
        return sum(l.get("qty",0) or 0 for l in lines if l.get("assembly")==asm)
    def _bom_count(lines,asm):
        return sum(1 for l in lines if l.get("assembly")==asm)

    assemblies = list(dict.fromkeys(r["assembly"] for r in cbom_proto))
    for ri,asm in enumerate(assemblies,2):
        bg=ROW_ALT if ri%2 else ROW_WHITE
        sb_sum  = _bom_sum(bom_lines,asm)
        p_sum   = _asm_sum(cbom_proto,asm)
        v1_sum  = _asm_sum(cbom_vl1,asm)
        v2_sum  = _asm_sum(cbom_vl2,asm)
        sb_cnt  = _bom_count(bom_lines,asm)
        p_cnt   = _asm_count(cbom_proto,asm)
        v1_cnt  = _asm_count(cbom_vl1,asm)
        v2_cnt  = _asm_count(cbom_vl2,asm)
        sum_same   = (abs(p_sum-sb_sum)<0.0001)
        count_same = (p_cnt==sb_cnt)
        for col,val in enumerate([asm,sb_sum,p_sum,v1_sum,v2_sum,
                                   sb_cnt,p_cnt,v1_cnt,v2_cnt,
                                   sum_same,count_same],1):
            c=_data_cell(ws,ri,col,val,bg=bg,
                         align="right" if isinstance(val,float) else
                               "center" if isinstance(val,bool) else "left")
        ws.row_dimensions[ri].height=15
    return ws

# ── Sum & Count Detailed ──────────────────────────────────────────────────────
def _build_sum_count_detailed(wb, cbom_proto, cbom_vl1, cbom_vl2, bom_lines):
    ws = wb.create_sheet("Sum & Count Detailed")
    ws.sheet_view.showGridLines = False
    hdrs = ["Assembly","Part Number","Scrub BOM Sum","Scrub BOM Count",
            "Proto Sum","Proto Count","VL-1 Sum","VL-1 Count",
            "VL-2 Sum","VL-2 Count","Is Sum Same","Is Count Same"]
    for col,h in enumerate(hdrs,1):
        _hdr_cell(ws,1,col,h)
        ws.column_dimensions[get_column_letter(col)].width=16 if col>2 else 22
    ws.row_dimensions[1].height=20

    # Group by (assembly, cpn)
    keys = list(dict.fromkeys((r["assembly"],r["cpn"]) for r in cbom_proto))
    bom_lookup = {(l["assembly"],l["cpn"]):l for l in bom_lines}

    for ri,(asm,cpn) in enumerate(keys,2):
        bg=ROW_ALT if ri%2 else ROW_WHITE
        bom_l = bom_lookup.get((asm,cpn),{})
        sb_sum  = bom_l.get("qty",0) or 0
        sb_cnt  = 1 if bom_l else 0

        def _vsum(vol,a,c):
            return sum(r.get("part_qty",0) or 0 for r in vol
                       if r.get("assembly")==a and r.get("cpn")==c)
        p_sum  = _vsum(cbom_proto,asm,cpn)
        v1_sum = _vsum(cbom_vl1,asm,cpn)
        v2_sum = _vsum(cbom_vl2,asm,cpn)
        p_cnt  = sum(1 for r in cbom_proto if r.get("assembly")==asm and r.get("cpn")==cpn)
        v1_cnt = sum(1 for r in cbom_vl1   if r.get("assembly")==asm and r.get("cpn")==cpn)
        v2_cnt = sum(1 for r in cbom_vl2   if r.get("assembly")==asm and r.get("cpn")==cpn)
        ss = abs(p_sum-sb_sum)<0.0001
        cs = p_cnt==sb_cnt

        for col,val in enumerate([asm,cpn,sb_sum,sb_cnt,p_sum,p_cnt,
                                   v1_sum,v1_cnt,v2_sum,v2_cnt,ss,cs],1):
            _data_cell(ws,ri,col,val,bg=bg,
                       align="right" if isinstance(val,(int,float)) and not isinstance(val,bool)
                             else "center" if isinstance(val,bool) else "left")
        ws.row_dimensions[ri].height=15
    return ws

# ── Revision History ─────────────────────────────────────────────────────────
def _build_revision_history(wb, project):
    ws = wb.create_sheet("Revision History")
    ws.sheet_view.showGridLines = False
    hdrs = [("Rev No",10),("Date",14),("Description",60),("Changed By",20),("Approved By",20)]
    for col,(h,w) in enumerate(hdrs,1):
        _hdr_cell(ws,1,col,h)
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.row_dimensions[1].height=20

    # Current rev
    _data_cell(ws,2,1, project.get("rev_no",0),   align="center")
    _data_cell(ws,2,2, datetime.utcnow().strftime("%d-%b-%Y"))
    _data_cell(ws,2,3, f'Initial CBOM generated by BuLLMQuote — {project.get("code","")}')
    _data_cell(ws,2,4, "BuLLMQuote System")
    _data_cell(ws,2,5, "")
    ws.row_dimensions[2].height=15
    return ws

# ── Price Control ─────────────────────────────────────────────────────────────
def _build_price_control(wb, cbom_vl2):
    ws = wb.create_sheet("Price Control")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1,column=1,value="Price Control").font=_font(bold=True,size=11)
    hdrs = [("Price Control",20),
            ("Sum of Ext Price (Conv.)",22),
            ("Sum of Ext Vol Cost (Conv.)",24),
            ("In %",10)]
    for col,(h,w) in enumerate(hdrs,1):
        _hdr_cell(ws,4,col,h)
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.row_dimensions[4].height=20

    pc_groups: Dict[str,Dict] = defaultdict(lambda:{"ext_price":0.0,"ext_vol":0.0})
    grand_ext_price=0.0; grand_ext_vol=0.0
    for r in cbom_vl2:
        pc = r.get("price_control","centum") or "centum"
        pc_groups[pc]["ext_price"] += r.get("ext_price_conv",0) or 0
        pc_groups[pc]["ext_vol"]   += r.get("ext_vol_price",0) or 0
        grand_ext_price += r.get("ext_price_conv",0) or 0
        grand_ext_vol   += r.get("ext_vol_price",0) or 0

    for ri,(pc,d) in enumerate(pc_groups.items(),5):
        bg=ROW_ALT if ri%2 else ROW_WHITE
        pct = d["ext_price"]/grand_ext_price if grand_ext_price else 0
        _data_cell(ws,ri,1,pc,bg=bg)
        _data_cell(ws,ri,2,d["ext_price"],bg=bg,align="right",num_fmt=NUM_PRICE)
        _data_cell(ws,ri,3,d["ext_vol"],  bg=bg,align="right",num_fmt=NUM_PRICE)
        _data_cell(ws,ri,4,pct,           bg=bg,align="right",num_fmt=NUM_PCT)
        ws.row_dimensions[ri].height=15

    tr=len(pc_groups)+5
    _hdr_cell(ws,tr,1,"Grand Total")
    _hdr_cell(ws,tr,2,grand_ext_price,align="right"); ws.cell(row=tr,column=2).number_format=NUM_PRICE
    _hdr_cell(ws,tr,3,grand_ext_vol,  align="right"); ws.cell(row=tr,column=3).number_format=NUM_PRICE
    _hdr_cell(ws,tr,4,1.0,            align="right"); ws.cell(row=tr,column=4).number_format=NUM_PCT
    return ws

# ── Summary ───────────────────────────────────────────────────────────────────
def _build_summary(wb, project, asm_summary_proto, asm_summary_vl1, asm_summary_vl2):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.cell(row=3,column=3,
            value=f'{project.get("code","")}, Rev-{project.get("rev_no",0)}, '
                  f'{datetime.utcnow().strftime("%d-%b-%Y")}').font=_font(bold=True)

    hdrs = ["Assembly","Rev No","Proto Qty","CBOM Cost Proto",
            "Volume-1","CBOM Cost VL-1","Volume-2","CBOM Cost VL-2","CBOM Ref No"]
    for col,h in enumerate(hdrs,3):
        _hdr_cell(ws,4,col,h)
        ws.column_dimensions[get_column_letter(col)].width = (
            20 if col in [3,11] else 14 if col in [4,7,9] else 18
        )

    assemblies = sorted(set(
        list(asm_summary_proto.keys())+
        list(asm_summary_vl1.keys())+
        list(asm_summary_vl2.keys())
    ))
    for ri,asm in enumerate(assemblies,5):
        bg=ROW_ALT if ri%2 else ROW_WHITE
        _data_cell(ws,ri,3,asm,bg=bg,bold=True)
        _data_cell(ws,ri,4,project.get("rev_no",0),bg=bg,align="center")
        _data_cell(ws,ri,5,project.get("proto_qty",10),bg=bg,align="right")
        _data_cell(ws,ri,6,asm_summary_proto.get(asm,{}).get("ext_price_sum",0),
                   bg=bg,align="right",num_fmt=NUM_PRICE)
        _data_cell(ws,ri,7,project.get("vl1_qty",300),bg=bg,align="right")
        _data_cell(ws,ri,8,asm_summary_vl1.get(asm,{}).get("ext_price_sum",0),
                   bg=bg,align="right",num_fmt=NUM_PRICE)
        _data_cell(ws,ri,9,project.get("vl2_qty",1000),bg=bg,align="right")
        _data_cell(ws,ri,10,asm_summary_vl2.get(asm,{}).get("ext_price_sum",0),
                   bg=bg,align="right",num_fmt=NUM_PRICE)
        _data_cell(ws,ri,11,project.get("code",""),bg=bg)
        ws.row_dimensions[ri].height=15
    return ws

# ── Batchwise Cashflow ────────────────────────────────────────────────────────
def _build_cashflow(wb, cbom_proto, cbom_vl1, cbom_vl2):
    ws = wb.create_sheet("Batchwise Cashflow")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1,column=1,value="Cashflow details").font=_font(bold=True,size=11)

    def _section(start_row, title, rows, vol_label):
        ws.cell(row=start_row,column=1,value=title).font=_font(bold=True)
        _hdr_cell(ws,start_row+1,1,"Supplier Type",bg=HDR_MED)
        _hdr_cell(ws,start_row+1,2,"Sum of Final Buy Value",bg=HDR_MED)
        _hdr_cell(ws,start_row+1,3,"Avg Payment Terms",bg=HDR_MED)
        for c in [1,2,3]:
            ws.column_dimensions[get_column_letter(c)].width=28

        pt_groups: Dict[str,float] = defaultdict(float)
        for r in rows:
            pt = r.get("payment_term","") or "Unknown"
            pt_groups[pt] += r.get("ext_vol_price",0) or 0

        ri_off=start_row+2; total=0.0
        for pt,val in sorted(pt_groups.items()):
            bg=ROW_ALT if ri_off%2 else ROW_WHITE
            _data_cell(ws,ri_off,1,pt,bg=bg)
            _data_cell(ws,ri_off,2,val,bg=bg,align="right",num_fmt=NUM_PRICE)
            total+=val; ri_off+=1
        _hdr_cell(ws,ri_off,1,"Grand Total")
        _hdr_cell(ws,ri_off,2,total,align="right"); ws.cell(row=ri_off,column=2).number_format=NUM_PRICE
        return ri_off+2

    next_row = _section(3,  "Payment term details for Proto",  cbom_proto, "Proto")
    next_row = _section(next_row, "Payment term details for VL-1", cbom_vl1,   "VL-1")
    _section(next_row, "Payment term details for VL-2", cbom_vl2,   "VL-2")
    return ws

# ── Excess Inventory ─────────────────────────────────────────────────────────
def _build_ex_inv(wb, sheet_name, cbom_rows, board_qty):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    hdrs = [("Part Number",20),("Description",36),("Commodity",14),
            ("Mfg Name",24),("Mfg Part Number",28),("Part Qty",10),
            ("Ext Vol",10),("Excess In Stk",13),("Excess After Demand",16),
            ("Shortage Qty",12),("Scrap",8),("Scrap Factor",11),
            ("Unit Cost",12),("Ext Vol Cost",13),
            ("Supplier",28),("MOQ",8)]
    for col,(h,w) in enumerate(hdrs,1):
        _hdr_cell(ws,17,col,h,wrap=True)
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.row_dimensions[17].height=30
    ws.freeze_panes="A18"

    # Total header
    ws.cell(row=12,column=1,value="Excess Cost (Conv.):").font=_font(bold=True)
    total_exc = sum(
        (r.get("unit_price_conv",0) or 0)*(r.get("excess_in_stk",0) or 0)
        for r in cbom_rows
    )
    ws.cell(row=13,column=1,value=round(total_exc,4)).number_format=NUM_PRICE

    for ri,r in enumerate(cbom_rows,18):
        bg=ROW_ALT if ri%2 else ROW_WHITE
        exc_stk = r.get("excess_in_stk",0) or 0
        short   = r.get("shortage_qty",0) or 0
        scrap   = (r.get("scrap_qty",0) or 0)
        sf      = r.get("scrap_factor",0) or 0
        up      = r.get("unit_price_conv",0) or 0
        ev      = r.get("ext_vol_qty",0) or 0
        ev_cost = up * ev

        flag_bg = RED_BG if short > 0 else (YELLOW_BG if exc_stk > 0 else bg)
        for col,val in enumerate([
            r.get("cpn",""),r.get("description",""),r.get("commodity",""),
            r.get("manufacturer",""),r.get("mpn",""),
            r.get("part_qty",0),ev,exc_stk,
            max(0,exc_stk - short),short,scrap,sf,
            up,ev_cost,r.get("supp_name",""),r.get("moq",0)
        ],1):
            nm = (NUM_PRICE if col in [13,14] else
                  NUM_QTY   if col in [6,7,8,9,10,11,16] else
                  NUM_PCT   if col==12 else None)
            _data_cell(ws,ri,col,val,
                       bg=flag_bg if col in [8,9,10] else bg,
                       align="right" if isinstance(val,(int,float)) else "left",
                       num_fmt=nm)
        ws.row_dimensions[ri].height=15
    return ws

# ══════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
def build_workbook(
    project:      Dict,
    bom_lines:    List[Dict],
    qw_resolved:  Dict,
    cbom:         Dict,         # {"PROTO":[], "VL1":[], "VL2":[]}
    nre_lines:    List[Dict],
) -> bytes:
    from services.cbom_engine import assembly_summary

    proto = cbom["PROTO"]
    vl1   = cbom["VL1"]
    vl2   = cbom["VL2"]

    asm_proto = assembly_summary(proto)
    asm_vl1   = assembly_summary(vl1)
    asm_vl2   = assembly_summary(vl2)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    # 1-3: CBOM sheets
    _build_cbom_sheet(wb,"CBOM Proto",    proto, project.get("proto_qty",10),  project, asm_proto)
    _build_cbom_sheet(wb,"CBOM VL-1",    vl1,   project.get("vl1_qty",300),   project, asm_vl1)
    _build_cbom_sheet(wb,"CBOM VL-2",    vl2,   project.get("vl2_qty",1000),  project, asm_vl2)

    # 4-6: BOM Matrix
    _build_bom_matrix(wb,"BOM MATRIX - Proto",proto,project.get("proto_qty",10))
    _build_bom_matrix(wb,"BOM MATRIX - VL1",  vl1,  project.get("vl1_qty",300))
    _build_bom_matrix(wb,"BOM MATRIX - VL2",  vl2,  project.get("vl2_qty",1000))

    # 7: Part Num vs Mfg Num
    _build_part_vs_mfg(wb, bom_lines, qw_resolved)

    # 8: NRE Charges
    _build_nre(wb, nre_lines)

    # 9: Missing Notes
    _build_missing_notes(wb, proto)

    # 10: Lead Time FG Wise
    _build_lead_time_fg(wb, proto)

    # 11-12: Sum & Count
    _build_sum_count(wb, proto, vl1, vl2, bom_lines)
    _build_sum_count_detailed(wb, proto, vl1, vl2, bom_lines)

    # 13: Revision History
    _build_revision_history(wb, project)

    # 14: Price Control
    _build_price_control(wb, vl2)

    # 15: Summary
    _build_summary(wb, project, asm_proto, asm_vl1, asm_vl2)

    # 16: Batchwise Cashflow
    _build_cashflow(wb, proto, vl1, vl2)

    # 17-19: Excess Inventory
    _build_ex_inv(wb, "Excess Inv Proto", proto, project.get("proto_qty",10))
    _build_ex_inv(wb, "Excess Inv VL-1",  vl1,   project.get("vl1_qty",300))
    _build_ex_inv(wb, "Excess Inv VL-2",  vl2,   project.get("vl2_qty",1000))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
